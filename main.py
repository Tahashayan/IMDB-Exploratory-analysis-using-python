import openpyxl
import requests
from bs4 import BeautifulSoup

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = "Top movies"
print(excel.sheetnames)
sheet.append(['Title', 'Year_Release', 'Genre', 'Rating', 'Votes'])

html_url = requests.get('https://www.imdb.com/search/title/?genres=Action&explore=genres&title_type=feature&ref_=ft_movie_0').text
soup = BeautifulSoup(html_url, 'html.parser')
movie_list = soup.find_all('div', class_='lister-item mode-advanced')
for movie in movie_list:
    movie_title = movie.h3.a.text
    movie_year = movie.find('span', class_='lister-item-year text-muted unbold').text
    movie_genre = movie.find('span', class_='genre').text
    movie_rating = movie.find_next('strong').text
    movie_votes = movie.find_next('span', {'name': 'nv'}).text
    print("\n")
    sheet.append([movie_title, movie_year, movie_genre, movie_rating, movie_votes])
excel.save('Idbm movies.xlsx')
